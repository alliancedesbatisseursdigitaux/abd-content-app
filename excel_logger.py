# excel_logger.py — Enregistrement des posts dans un fichier Excel
# Module pour suivre tous les posts générés et publiés

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime

# Couleurs ABD
VERT_ABD = "00C853"
VIOLET_ABD = "6B2FB5"

def log_to_excel(date, jour, reseau, pilier, sujet, texte, image_path, statut, url_post="", notes=""):
    """Ajoute une ligne au fichier Excel de suivi"""
    
    excel_path = "output/posts_log.xlsx"
    
    # Créer le dossier output s'il n'existe pas
    os.makedirs("output", exist_ok=True)
    
    # Créer ou charger le fichier Excel
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Suivi des posts"
        
        # En-têtes
        headers = [
            "Date", "Jour", "Réseau social", "Pilier", "Sujet",
            "Texte (extrait)", "Chemin image", "Statut", "URL du post",
            "Likes", "Commentaires", "Partages", "Vues", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=VIOLET_ABD, end_color=VIOLET_ABD, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 10
        ws.column_dimensions['N'].width = 30
    
    # Ajouter la ligne
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=date)
    ws.cell(row=row, column=2, value=jour)
    ws.cell(row=row, column=3, value=reseau)
    ws.cell(row=row, column=4, value=pilier)
    ws.cell(row=row, column=5, value=sujet)
    ws.cell(row=row, column=6, value=texte)
    ws.cell(row=row, column=7, value=image_path)
    ws.cell(row=row, column=8, value=statut)
    ws.cell(row=row, column=9, value=url_post)
    ws.cell(row=row, column=10, value="")  # Likes (à remplir manuellement)
    ws.cell(row=row, column=11, value="")  # Commentaires
    ws.cell(row=row, column=12, value="")  # Partages
    ws.cell(row=row, column=13, value="")  # Vues
    ws.cell(row=row, column=14, value=notes)
    
    # Colorer la ligne selon le statut
    if statut == "Publié":
        fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    elif statut == "Échec":
        fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    else:
        fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    for col in range(1, 15):
        ws.cell(row=row, column=col).fill = fill
    
    # Sauvegarder
    wb.save(excel_path)
    return True
